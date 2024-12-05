from functools import reduce
import openpyxl as pyx
from openpyxl.styles import Font, Alignment, Color
from openpyxl.styles.borders import Border, BORDER_THIN, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from SchichtplanUtils import getLettersForRef


def writeData(data, file_name, sheet_names):
    sheet_name = "Zusammenfassung"
    schichtplan_wb = pyx.load_workbook(file_name)

    new_sheet = schichtplan_wb.create_sheet(sheet_name)
    new_sheet.title = sheet_name

    schichtplan_sh = schichtplan_wb[sheet_name]

    schichtplan_sh.sheet_view.zoomScale = 70

    itnl_row_count = 1

    for data_idx, i in enumerate(data):

        if data_idx > 0:
            itnl_row_count+=2

        #Append Title of Dict
        title_row_idx = itnl_row_count

        title_cell_obj = schichtplan_sh.cell(row=title_row_idx, column=1)
        title_cell_obj.value = (i["name"])
        title_cell_obj.font = Font(bold=True,size=20) #font setzen
        title_cell_obj.alignment = Alignment(horizontal='center', vertical='center')
        title_cell_obj.border = Border(
                                        left=Side(border_style=BORDER_THIN, color='00000000'),
                                        right=Side(border_style=BORDER_THIN, color='00000000'),
                                        top=Side(border_style=BORDER_THIN, color='00000000'),
                                        bottom=Side(border_style=BORDER_THIN, color='00000000')
                                    )
        
        #Add KW title
        kw_title_cell_obj = schichtplan_sh.cell(row=title_row_idx, column=2)
        kw_title_cell_obj.value = "KW"
        kw_title_cell_obj.font = Font(bold=True,size=20) #font setzen
        kw_title_cell_obj.alignment = Alignment(horizontal='center', vertical='center')
        
        itnl_row_count+=1

        #Append Header of Table
        header_row_idx = title_row_idx + 1
        for name_idx, name in enumerate(sheet_names):
            header_cell_obj = schichtplan_sh.cell(row=header_row_idx, column=name_idx+1)
            if name_idx > 0:
                if name[2:4].startswith("0"):
                    header_cell_obj.value = name[2:4].replace('0','')
                if not name[2:4].startswith("0"):
                    header_cell_obj.value = name[2:4]
            if name_idx == 0:
                header_cell_obj.value = name
            header_cell_obj.font = Font(bold=True, size=15, color=Color("FFFFFFFF")) #font setzen
        
        #Append Sum Header of Table
        sum_title_cell_obj = schichtplan_sh.cell(row=header_row_idx, column=(len(sheet_names)+1))
        sum_title_cell_obj.value = "Summe"
        sum_title_cell_obj.font = Font(bold=True, size=15, color=Color("FFFFFFFF")) #font setzen


        itnl_row_count+=1

        #Append Data of Table
        item_idx = 0
        for k, v in dict(sorted(data[data_idx]["dic"].items())).items():
            # name in die erste spalte eintragen
            name_cell_obj = schichtplan_sh.cell(row=itnl_row_count, column=1)
            name_cell_obj.value = k
            name_cell_obj.font = Font(size=16, color=Color(rgb=v["farbe"]), bold=v["bold"])

            # die tage in die richtige spalte eintragen
            for sheet_idx, sheet in enumerate(v["sheet"]):
                appendInIdx = sheet_names.index(sheet)+1
                day_cell_obj = schichtplan_sh.cell(row=itnl_row_count, column= appendInIdx)
                day_cell_obj.value = v["tage"][sheet_idx]
                day_cell_obj.font = Font(size=15)
            
            # summierten wert hinzufügen
            sum_cell_obj = schichtplan_sh.cell(row=itnl_row_count, column=len(sheet_names)+1)
            sumFunc = lambda x, y: x+y
            sum_cell_obj.value = reduce(sumFunc, v["tage"])
            sum_cell_obj.font = Font(size=15)

            item_idx+=1
            itnl_row_count+=1
        
        #Format as Table

        tableRef = "A" + str(header_row_idx) + ":" + getLettersForRef(sheet_names) + str(itnl_row_count-1)
        table = Table(displayName=i["name"], ref=tableRef)
        style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        schichtplan_sh.add_table(table)

    schichtplan_wb.save(file_name)