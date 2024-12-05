import openpyxl as pyx
import string

def getData(file_name):
    schichtplan_wb = pyx.load_workbook(file_name)
    sheet_name = "Zusammenfassung"

    if sheet_name in schichtplan_wb.sheetnames:
        del schichtplan_wb[sheet_name]
        schichtplan_wb.save(file_name)

    sheetNames=schichtplan_wb.sheetnames

    mitarbeiterZeile = 5
    mitarbeiterSpalte = 2

    readWochentage = 7

    krank_dic = {}
    urlaub_dic = {}
    at_dic = {}
    stundenabbau_dic = {}

    dicList = [{"name": "Krank", "dic": krank_dic},{"name": "Urlaub", "dic": urlaub_dic},{"name": "AT", "dic": at_dic},{"name": "Stundenabbau", "dic": stundenabbau_dic}]

    for l in sheetNames:
        schichtplan_sh = schichtplan_wb[l]
        # initialisieren der daten für jeweiliges sheet
        for i in range(mitarbeiterZeile, schichtplan_sh.max_row+1): 
            for j in range(2, readWochentage+1):
                zellen_obj = schichtplan_sh.cell(row=i, column=j)
                # Mitarbeiter einlesen und eintragen, wenn nicht vorhanden
                if j == 2:
                    if zellen_obj.value: 
                        if zellen_obj.value.strip() not in ["Frühschicht", "Spätschicht", "Nachtschicht", "Anzahl der MA"]:
                            # Für den ersten Durchlauf
                            # mitarbeiter in dicts anlegen
                            for dic in dicList:
                                if not zellen_obj.value.strip() in dic["dic"]:
                                    if zellen_obj.font.color and 'rgb' in zellen_obj.font.color.__dict__:
                                        dic["dic"][zellen_obj.value.strip()] = {"sheet": list(sheetNames),"tage": ([0] * len(sheetNames)), "farbe": str(zellen_obj.font.color.rgb), "bold": zellen_obj.font.bold}
                                    else:
                                        dic["dic"][zellen_obj.value.strip()] = {"sheet": list(sheetNames),"tage": ([0] * len(sheetNames)), "farbe": "00000000", "bold": False}
                # Wochentage einlesen und prüfen
                if j >= 3:
                    mitarbeiterZelle_obj = schichtplan_sh.cell(row=i, column=mitarbeiterSpalte)
                    if mitarbeiterZelle_obj.value:
                        if mitarbeiterZelle_obj.value.strip() not in ["Frühschicht", "Spätschicht", "Nachtschicht", "Anzahl der MA", "N.N."]:
                            sheetIndex = dicList[0]["dic"][mitarbeiterZelle_obj.value.strip()]["sheet"].index(l)
                            if zellen_obj.value:
                                if str(zellen_obj.value).strip() == "U":
                                    dicList[1]["dic"][mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] = urlaub_dic[mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] + 1
                                if str(zellen_obj.value).strip() == "AT":
                                    dicList[2]["dic"][mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] = at_dic[mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] + 1
                                if str(zellen_obj.value).strip() == "Krank":
                                    dicList[0]["dic"][mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] = krank_dic[mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] + 1
                                if str(zellen_obj.value).strip() == "StAb":
                                    dicList[3]["dic"][mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] = stundenabbau_dic[mitarbeiterZelle_obj.value.strip()]["tage"][sheetIndex] + 1
    return (dicList, sheetNames)